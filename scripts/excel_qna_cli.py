#!/usr/bin/env python3

"""
excel_qna_cli.py

Quick testing CLI to answer questions listed in an Excel sheet.

Expected columns (header row):
  - Team
  - Questions
  - Fund
  - Answer
  - Answer Context

For each row with a non-empty Questions cell, the tool calls the
shared QA engine (qa_core.answer_question) to produce an answer and
context/citations. It writes the answer into the Answer column and a
compact textual context into the Answer Context column.

Usage examples:
  python excel_qna_cli.py input.xlsx
  python excel_qna_cli.py input.xlsx -o output.xlsx --sheet Sheet1 --overwrite
  python excel_qna_cli.py input.xlsx --length medium --mode both --k 6

Note: The default LLM client uses the custom CompletionsClient. Ensure
the necessary environment variables are set (see backend/llm/completions_client.py).
"""

from __future__ import annotations

import argparse
import math
import os
import re
from collections import Counter
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Tuple

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from backend.llm.completions_client import CompletionsClient, get_openai_completion
from backend.answering.qa_engine import answer_question


def _find_column_indexes(
    ws: Worksheet,
    header_row: int,
    names: Dict[str, Iterable[str] | str],
) -> Dict[str, int]:
    """Return 0-based column indexes for required columns by header name.

    names keys: team, question, fund, answer, context, approved, score
    """
    headers: Dict[str, int] = {}
    last_col = ws.max_column or 0
    lookup: Dict[str, str] = {}
    for key, raw in names.items():
        if not raw:
            continue
        if isinstance(raw, (list, tuple, set)):
            for name in raw:
                if name:
                    lookup[str(name).strip().lower()] = key
        else:
            lookup[str(raw).strip().lower()] = key
    for col in range(1, last_col + 1):
        val = ws.cell(row=header_row, column=col).value
        if val is None:
            continue
        key = str(val).strip().lower()
        if key in lookup and lookup[key] not in headers:
            headers[lookup[key]] = col - 1
    missing = [k for k in ("question", "answer", "context") if k not in headers]
    if missing:
        raise SystemExit(
            f"Missing required columns in header row {header_row}: {', '.join(missing)}"
        )
    # Optional columns
    headers.setdefault("team", None)  # type: ignore[arg-type]
    headers.setdefault("fund", None)  # type: ignore[arg-type]
    headers.setdefault("approved", None)  # type: ignore[arg-type]
    headers.setdefault("score", None)  # type: ignore[arg-type]
    return headers


def _format_context(
    comments: List[Tuple[str, str, str, float, str]],
) -> str:
    """Build a compact, single-cell context summary from comments.

    Each comment is (label, source_name, snippet, score, date_str)
    """
    parts: List[str] = []
    for lbl, src, snippet, score, date_str in comments:
        snippet_one = (snippet or "").replace("\n", " ")
        if len(snippet_one) > 240:
            snippet_one = snippet_one[:237] + "…"
        date_piece = f" {date_str}" if date_str and date_str != "unknown" else ""
        parts.append(f"{lbl}) {src}:{date_piece} {snippet_one}")
    return "\n".join(parts)


_WORD_RE = re.compile(r"[A-Za-z0-9']+")


def _text_cosine_similarity(a: str, b: str) -> float:
    """Compute cosine similarity between two strings using token frequency."""
    if not a or not b:
        return 0.0
    vec_a = Counter(_WORD_RE.findall(a.lower()))
    vec_b = Counter(_WORD_RE.findall(b.lower()))
    if not vec_a or not vec_b:
        return 0.0
    dot = sum(vec_a[tok] * vec_b[tok] for tok in set(vec_a) & set(vec_b))
    norm_a = math.sqrt(sum(count * count for count in vec_a.values()))
    norm_b = math.sqrt(sum(count * count for count in vec_b.values()))
    if norm_a == 0.0 or norm_b == 0.0:
        return 0.0
    return dot / (norm_a * norm_b)


class OpenAIClient:
    def __init__(self, model: str):
        self.model = model

    def get_completion(self, prompt: str, json_output: bool = False, **kwargs):
        messages = kwargs.get("messages")
        if messages:
            prompt = messages[-1]["prompt"]
        return get_openai_completion(prompt, self.model, json_output=json_output)


def run(
    input_path: Path,
    output_path: Optional[Path],
    sheet_name: Optional[str],
    header_row: int,
    overwrite: bool,
    mode: str,
    k: int,
    min_conf: float,
    length: Optional[str],
    approx_words: Optional[int],
    framework: str,
    model_name: str,
) -> Path:
    wb = load_workbook(str(input_path))
    ws: Worksheet
    if sheet_name:
        if sheet_name not in wb.sheetnames:
            raise SystemExit(
                f"Worksheet '{sheet_name}' not found in {input_path.name}."
            )
        ws = wb[sheet_name]
    else:
        ws = wb.active

    col_names = {
        "team": "Team",
        "question": "Questions",
        "fund": "Fund",
        "answer": "Answer",
        "context": "Answer Context",
        "approved": ("Approved Answer", "Approved Answers", "Approved Answer Text"),
        "score": ("Score", "Similarity Score"),
    }
    idx = _find_column_indexes(ws, header_row, col_names)

    approved_idx = idx.get("approved")
    score_idx = idx.get("score")
    if approved_idx is not None and score_idx is None:
        print(
            "Warning: Found an 'Approved Answer' column but no 'Score' column; cosine scores will be skipped."
        )
    if score_idx is not None and approved_idx is None:
        print(
            "Warning: Found a 'Score' column but no 'Approved Answer' column; cosine scores will be skipped."
        )

    if framework == "openai":
        client = OpenAIClient(model=model_name)
    else:
        client = CompletionsClient(model=model_name)

    start_row = header_row + 1
    end_row = ws.max_row or start_row - 1

    answered = 0
    skipped = 0
    errors = 0

    for r in range(start_row, end_row + 1):
        q_cell = ws.cell(row=r, column=idx["question"] + 1)
        a_cell = ws.cell(row=r, column=idx["answer"] + 1)
        c_cell = ws.cell(row=r, column=idx["context"] + 1)
        approved_cell = (
            ws.cell(row=r, column=approved_idx + 1) if approved_idx is not None else None
        )
        score_cell = (
            ws.cell(row=r, column=score_idx + 1) if score_idx is not None else None
        )

        q_raw = q_cell.value
        q = (str(q_raw).strip() if q_raw is not None else "")
        if not q:
            skipped += 1
            continue

        if not overwrite and (a_cell.value not in (None, "") or c_cell.value not in (None, "")):
            skipped += 1
            continue

        # Resolve Fund used for this row and print it
        fund_val = None
        fund_has_column = idx.get("fund") is not None
        if fund_has_column:
            f_cell = ws.cell(row=r, column=idx["fund"] + 1)
            if f_cell.value not in (None, ""):
                fund_val = str(f_cell.value).strip()
        if not fund_has_column:
            print(f"[Row {r}] No 'Fund' column; using fund=None")
        elif fund_val:
            print(f"[Row {r}] Using fund='{fund_val}'")
        else:
            print(f"[Row {r}] Blank 'Fund' cell; using fund=None")

        try:
            ans, comments = answer_question(
                q=q,
                mode=mode,
                fund=fund_val,
                k=k,
                length=length,
                approx_words=approx_words,
                min_confidence=min_conf,
                llm=client,
            )

            context_text = _format_context(comments)

            a_cell.value = ans
            c_cell.value = context_text
            if approved_cell is not None and score_cell is not None:
                approved_text = approved_cell.value
                approved_str = (
                    str(approved_text).strip() if approved_text not in (None, "") else ""
                )
                ans_str = ans if isinstance(ans, str) else str(ans or "")
                if approved_str and ans_str:
                    similarity = _text_cosine_similarity(ans_str, approved_str)
                    score_cell.value = round(similarity, 4)
            # make wrap
            try:
                a_cell.alignment = a_cell.alignment.copy(wrap_text=True)
                c_cell.alignment = c_cell.alignment.copy(wrap_text=True)
            except Exception:
                pass
            answered += 1
        except Exception as e:
            c_cell.value = f"Error: {e}"
            try:
                c_cell.alignment = c_cell.alignment.copy(wrap_text=True)
            except Exception:
                pass
            errors += 1

    out_path = output_path or input_path.with_name(f"{input_path.stem}_answered{input_path.suffix}")
    wb.save(str(out_path))
    wb.close()

    print(
        f"Processed rows {start_row}-{end_row}: answered={answered}, skipped={skipped}, errors={errors}. Saved to {out_path}"
    )
    return out_path


def main() -> None:
    ap = argparse.ArgumentParser(
        description="Answer questions from an Excel sheet and fill Answer/Answer Context columns.",
    )
    ap.add_argument("xlsx", help="Path to the input Excel file")
    ap.add_argument("-o", "--out", help="Path to output Excel; defaults to *_answered.xlsx")
    ap.add_argument("--sheet", help="Worksheet name (default: active sheet)")
    ap.add_argument("--header-row", type=int, default=1, help="Header row number (1-based)")
    ap.add_argument("--overwrite", action="store_true", help="Overwrite existing Answer/Context cells")
    ap.add_argument(
        "--mode",
        choices=["both", "question", "answer", "blend"],
        default="both",
        help="Vector search mode",
    )
    ap.add_argument("-k", type=int, default=20, help="Number of snippets to retrieve")
    ap.add_argument("--min-confidence", type=float, default=0.0, help="Minimum cosine to accept a snippet")
    ap.add_argument(
        "--length",
        choices=["short", "medium", "long", "auto"],
        default="long",
        help="Answer length preset",
    )
    ap.add_argument("--approx-words", type=int, help="Override target word count")
    ap.add_argument(
        "--framework",
        choices=["aladdin", "openai"],
        default=os.getenv("ANSWER_FRAMEWORK", "aladdin"),
        help="Backend framework for completions (defaults to 'aladdin').",
    )
    ap.add_argument(
        "--model",
        default=os.getenv("OPENAI_MODEL", "o3-2025-04-16_research"),
        help="Model name for the chosen framework (defaults to Streamlit's).",
    )

    args = ap.parse_args()

    in_path = Path(args.xlsx).expanduser().resolve()
    if not in_path.exists():
        raise SystemExit(f"Input Excel not found: {in_path}")

    out_path = Path(args.out).expanduser().resolve() if args.out else None

    run(
        input_path=in_path,
        output_path=out_path,
        sheet_name=args.sheet,
        header_row=args.header_row,
        overwrite=args.overwrite,
        mode=args.mode,
        k=args.k,
        min_conf=args.min_confidence,
        length=args.length,
        approx_words=args.approx_words,
        framework=args.framework,
        model_name=args.model,
    )


if __name__ == "__main__":
    main()
