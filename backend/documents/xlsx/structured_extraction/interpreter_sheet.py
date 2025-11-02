from __future__ import annotations

"""Lightweight wrappers used by legacy sheet interpretation workflows."""

from pathlib import Path
from typing import Dict, List

from openpyxl import load_workbook


def collect_non_empty_cells(path: str | Path) -> List[Dict[str, object]]:
    """Return a list of non-empty cells for the given workbook.

    Each item includes sheet name, row, column, and the cell value. This mirrors the
    behaviour expected by the legacy utilities that only need lightweight metadata.
    """

    workbook = load_workbook(filename=str(path), data_only=True)
    cells: List[Dict[str, object]] = []
    for sheet in workbook.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value not in (None, ""):
                    cells.append(
                        {
                            "sheet": sheet.title,
                            "row": cell.row,
                            "column": cell.column,
                            "value": cell.value,
                        }
                    )
    return cells


# Quick CLI helper:
# if __name__ == "__main__":
#     from pprint import pprint
#     pprint(collect_non_empty_cells("samples/workbook.xlsx")[:10])
