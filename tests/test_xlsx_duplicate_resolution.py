
import sys
import types

import pytest

class _DummyNLP:
    pipe_names = []
    def add_pipe(self, *a, **k):
        pass

def _load(name):
    return _DummyNLP()

def _blank(name):
    return _DummyNLP()
import pathlib
sys.path.insert(0, str(pathlib.Path(__file__).resolve().parent.parent))

sys.modules['spacy'] = types.SimpleNamespace(load=_load, blank=_blank)

pytest.importorskip("openpyxl")
try:
    import openpyxl.styles  # noqa: F401
except ModuleNotFoundError:
    pytest.skip(
        "openpyxl with styles support is required for duplicate resolution tests",
        allow_module_level=True,
    )

import openpyxl

if not hasattr(openpyxl, "Workbook"):
    pytest.skip(
        "openpyxl Workbook support is required for duplicate resolution tests",
        allow_module_level=True,
    )
from backend.documents.xlsx import slot_finder as finder
finder._spacy_is_question_or_imperative = lambda text: True


def test_resolves_duplicate_answer_cells(tmp_path, monkeypatch):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "First?"
    ws["A2"] = "Second?"
    in_path = tmp_path / "in.xlsx"
    wb.save(in_path)

    # Ensure framework permits LLM calls
    monkeypatch.setenv("OPENAI_API_KEY", "key")
    monkeypatch.setattr(finder, "FRAMEWORK", "openai")

    def fake_call_llm(prompt_file, payload, *, model):
        if prompt_file == "xlsx_workbook_answer_slot.txt":
            return {"sheet": "Sheet1", "answer_cell": "B1"}
        if prompt_file == "xlsx_resolve_duplicate_slots.txt":
            qs = payload["questions"]
            return [
                {
                    "question_sheet": qs[0]["question_sheet"],
                    "question_cell": qs[0]["question_cell"],
                    "sheet": "Sheet1",
                    "answer_cell": "B1",
                },
                {
                    "question_sheet": qs[1]["question_sheet"],
                    "question_cell": qs[1]["question_cell"],
                    "sheet": "Sheet1",
                    "answer_cell": "B2",
                },
            ]
        return {}

    monkeypatch.setattr(finder, "_call_llm", fake_call_llm)

    schema = finder.extract_schema_from_xlsx(str(in_path), debug=False)
    mapping = {(s["question_cell"], s["answer_cell"]) for s in schema}
    assert mapping == {("A1", "B1"), ("A2", "B2")}
