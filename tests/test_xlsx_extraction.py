import openpyxl
from backend.rfp_xlsx_slot_finder import extract_slots_from_xlsx, extract_schema_from_xlsx
from backend.rfp_xlsx_apply_answers import write_excel_answers


def test_extract_slots_from_xlsx(tmp_path):
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Sheet1"
    c = ws1["A1"]
    c.value = "Question?"
    c.font = openpyxl.styles.Font(color="FFFF0000", bold=True)
    c.fill = openpyxl.styles.PatternFill("solid", fgColor="FFFFFF00")
    ws2 = wb.create_sheet("Data")
    ws2["B2"] = "Answer"

    path = tmp_path / "sample.xlsx"
    wb.save(path)

    result = extract_slots_from_xlsx(str(path))
    assert result["doc_type"] == "xlsx"
    sheets = {s["name"]: s for s in result["sheets"]}
    assert set(sheets.keys()) == {"Sheet1", "Data"}

    sheet1_cells = {cell["address"]: cell for cell in sheets["Sheet1"]["cells"]}
    assert sheet1_cells["A1"]["value"] == "Question?"
    assert sheet1_cells["A1"]["bold"] is True
    assert sheet1_cells["A1"]["font_color"].upper().endswith("FF0000")
    data_cells = {cell["address"]: cell for cell in sheets["Data"]["cells"]}
    assert data_cells["B2"]["value"] == "Answer"


def test_question_without_answer_slot(tmp_path, capsys):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "What is your name?"
    ws["B1"] = "n/a"  # not blank
    in_path = tmp_path / "in.xlsx"
    wb.save(in_path)

    schema = extract_schema_from_xlsx(str(in_path), debug=False)
    assert schema and schema[0]["question_text"].startswith("What is")
    assert schema[0]["answer_cell"] is None

    out_path = tmp_path / "out.xlsx"
    res = write_excel_answers(schema, ["Alice"], str(in_path), str(out_path))
    assert res["applied"] == 0
    assert res["skipped"] == 1
    captured = capsys.readouterr()
    assert "no answer slot" in captured.out.lower()


def test_comment_formats_citations(tmp_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "Question?"
    in_path = tmp_path / "in.xlsx"
    wb.save(in_path)

    schema = [
        {
            "sheet": "Sheet1",
            "question_cell": "A1",
            "answer_cell": "B1",
            "question_text": "Question?",
        }
    ]
    answers = [
        {
            "text": "Ans [1] [2]",
            "citations": {
                1: {"text": "First", "source_file": "file1.txt"},
                2: {"text": "Second", "source_file": "file2.txt"},
            },
        }
    ]
    out_path = tmp_path / "out.xlsx"
    comments_path = tmp_path / "comments.docx"
    write_excel_answers(
        schema,
        answers,
        str(in_path),
        str(out_path),
        comments_docx_path=str(comments_path),
    )

    wb2 = openpyxl.load_workbook(out_path)
    c = wb2["Sheet1"]["B1"]
    assert c.comment is None
    assert c.value == "Ans"

    import docx
    from word_comments import ensure_comments_part

    doc = docx.Document(comments_path)
    part = ensure_comments_part(doc)
    xml = part._element.xml
    assert "First" in xml
    assert "Second" in xml
    assert "Source Text" in xml
    assert "Source File" in xml
    assert xml.index("Source File") < xml.index("Source Text")
    assert "file1.txt" in xml
    assert "file2.txt" in xml
    assert "<w:b/>" in xml


def test_preserves_pure_citation_answer(tmp_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "Question?"
    in_path = tmp_path / "in.xlsx"
    wb.save(in_path)

    schema = [
        {
            "sheet": "Sheet1",
            "question_cell": "A1",
            "answer_cell": "B1",
            "question_text": "Question?",
        }
    ]

    calls = {"count": 0}

    def generator(q):
        calls["count"] += 1
        return "Retry answer"

    answers = ["[1]"]
    out_path = tmp_path / "out.xlsx"
    write_excel_answers(
        schema, answers, str(in_path), str(out_path), generator=generator
    )

    assert calls["count"] == 0
    wb2 = openpyxl.load_workbook(out_path)
    c = wb2["Sheet1"]["B1"]
    assert c.value == "[1]"


def test_default_comments_docx_path(tmp_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "Question?"
    in_path = tmp_path / "in.xlsx"
    wb.save(in_path)

    schema = [
        {
            "sheet": "Sheet1",
            "question_cell": "A1",
            "answer_cell": "B1",
            "question_text": "Question?",
        }
    ]
    answers = [{"text": "Ans [1]", "citations": {1: "Snippet"}}]
    out_path = tmp_path / "out.xlsx"
    write_excel_answers(schema, answers, str(in_path), str(out_path))

    comments_path = tmp_path / "out_comments.docx"
    assert comments_path.exists()

    wb2 = openpyxl.load_workbook(out_path)
    c = wb2["Sheet1"]["B1"]
    assert c.comment is None
    assert c.value == "Ans"
