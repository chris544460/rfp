from __future__ import annotations

import json
import os
import re
from typing import Any, Dict, List, Optional, Tuple

from concurrent.futures import ThreadPoolExecutor, as_completed

import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Color
from openpyxl.utils import get_column_letter
import spacy


# Framework and model selection
FRAMEWORK = os.getenv("ANSWER_FRAMEWORK", "aladdin")

# Spreadsheet questions can benefit from a larger reasoning model.  Allow the
# Excel slot finder to use a dedicated model via ``OPENAI_MODEL_EXCEL_UNDERSTANDING``
# while leaving ``OPENAI_MODEL`` (used by other modules) at its usual default of
# ``gpt-4.1-nano-2025-04-14_research``.
EXCEL_MODEL = os.getenv(
    "OPENAI_MODEL_EXCEL_UNDERSTANDING", "o3-2025-04-16_research"
)

# Load spaCy model once at import time.  We prefer ``en_core_web_sm`` but fall
# back to a blank English pipeline if the model is unavailable.  The blank
# model still provides tokenization which is sufficient for our simple checks.
try:  # pragma: no cover - exercised implicitly during import
    _NLP = spacy.load("en_core_web_sm")
except Exception:  # pragma: no cover - missing model
    _NLP = spacy.blank("en")

# ``doc.sents`` requires sentence boundaries.  ``en_core_web_sm`` provides a
# dependency parser which sets them automatically.  The blank fallback pipeline
# only tokenizes text, so we add a simple ``sentencizer`` component that marks
# sentence starts from punctuation.  This mirrors spaCy's recommended fix for
# ``E030: sentence boundaries unset``.
if "parser" not in _NLP.pipe_names and "senter" not in _NLP.pipe_names:
    _NLP.add_pipe("sentencizer")


def _color_to_hex(color: Optional[Color]) -> Optional[str]:
    """Convert an openpyxl Color to a RGB hex string.

    openpyxl colors may include an alpha channel (ARGB).  This helper
    normalizes them to ``RRGGBB`` if possible.  If the color is theme- or
    indexed-based, ``None`` is returned.
    """
    if color is None:
        return None
    if getattr(color, "type", None) != "rgb":
        return None
    rgb = getattr(color, "rgb", None)
    if not isinstance(rgb, str):
        return None
    # ``rgb`` is usually ``AARRGGBB``; drop the alpha channel if present
    if len(rgb) == 8:
        rgb = rgb[2:]
    return rgb


# Basic question words used by the spaCy based detector.  These cover common
# English interrogatives.
QUESTION_WORDS = {
    "who",
    "what",
    "when",
    "where",
    "why",
    "how",
    "which",
}


def _spacy_is_question_or_imperative(text: str) -> bool:
    """Use spaCy to flag interrogative or imperative sentences.

    Each sentence is checked and considered a question if it ends with ``?`` or
    contains a question word.  Imperatives are detected when the sentence root
    has ``Mood=Imp`` or the first token is a bare verb (``tag_`` == ``VB``).
    """

    doc = _NLP(text)
    # ``doc.sents`` raises ``E030`` if no component sets sentence boundaries.
    # ``has_annotation("SENT_START")`` lets us cheaply check and fall back to the
    # whole doc when boundaries are missing.
    sentences = list(doc.sents) if doc.has_annotation("SENT_START") else [doc]
    for sent in sentences:
        sent_text = sent.text.strip()
        if not sent_text:
            continue
        if sent_text.endswith("?"):
            return True
        if any(tok.lower_ in QUESTION_WORDS for tok in sent):
            return True
        root = sent.root
        if "Imp" in root.morph.get("Mood"):
            return True
        first = sent[0]
        if root.tag_ == "VB" and first is root:
            return True
    return False


def _cell_info(cell) -> Dict[str, Any]:
    """Serialize a worksheet cell with basic formatting."""

    return {
        "address": cell.coordinate,
        "row": cell.row,
        "column": get_column_letter(cell.column),
        "value": None if cell.value is None else str(cell.value),
        "font_color": _color_to_hex(cell.font.color),
        "bold": bool(cell.font.bold),
        "italic": bool(cell.font.italic),
        "bg_color": _color_to_hex(cell.fill.start_color),
        "border": {
            "left": cell.border.left.style,
            "right": cell.border.right.style,
            "top": cell.border.top.style,
            "bottom": cell.border.bottom.style,
        },
    }


def _row_context(ws, row_idx: int) -> List[Dict[str, Any]]:
    """Return serialized cells for the entire ``row_idx`` of ``ws``."""

    return [_cell_info(ws.cell(row_idx, c)) for c in range(1, ws.max_column + 1)]


def _rect_context(ws, row_idx: int, col_idx: int, size: int = 10) -> List[Dict[str, Any]]:
    """Return serialized cells for a ``size``×``size`` box around (row, col).

    The box is clipped to the worksheet bounds when the question cell is near an
    edge so that we always return at most ``size``×``size`` cells.
    """

    half = size // 2
    start_row = max(1, row_idx - half)
    end_row = start_row + size - 1
    if end_row > ws.max_row:
        end_row = ws.max_row
        start_row = max(1, end_row - size + 1)

    start_col = max(1, col_idx - half)
    end_col = start_col + size - 1
    if end_col > ws.max_column:
        end_col = ws.max_column
        start_col = max(1, end_col - size + 1)

    rect: List[Dict[str, Any]] = []
    for r in range(start_row, end_row + 1):
        for c in range(start_col, end_col + 1):
            rect.append(_cell_info(ws.cell(r, c)))
    return rect


def _sheet_to_json(ws) -> Dict[str, Any]:
    """Serialize an entire worksheet to JSON-friendly structures."""

    cells: List[Dict[str, Any]] = []
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            cells.append(_cell_info(ws.cell(r, c)))
    return {"name": ws.title, "max_row": ws.max_row, "max_col": ws.max_column, "cells": cells}


def _question_context(ws, cell) -> Dict[str, Any]:
    """Build the question context used as LLM hints."""

    return {
        "question": str(cell.value).strip(),
        "cell": cell.coordinate,
        "row": _row_context(ws, cell.row),
        "rect": _rect_context(ws, cell.row, cell.column),
    }


def _llm_choose_answer_slot(
    question_ctx: Dict[str, Any],
    sheets: Dict[str, Any],
    *,
    model: str,
    debug: bool = False,
) -> Tuple[Optional[str], Optional[str]]:
    """Invoke the LLM to choose the answer sheet and cell.

    Parameters
    ----------
    question_ctx:
        Context dictionary describing the question cell.
    sheets:
        Mapping of sheet name to a JSON serialisation of the worksheet.  The
        structure mirrors the output of :func:`_sheet_to_json`.
    model:
        Name of the LLM model to invoke.

    Returns
    -------
    (sheet_name, cell_address):
        The selected sheet and cell address for the answer.  ``(None, None)`` is
        returned if the model cannot pick a location or the environment lacks
        credentials for the selected framework.
    """

    if FRAMEWORK == "openai":
        if not os.getenv("OPENAI_API_KEY"):
            if debug:
                print("  OPENAI_API_KEY not set for openai; skipping LLM call")
            return None, None
    elif FRAMEWORK == "aladdin":
        required = [
            "aladdin_studio_api_key",
            "defaultWebServer",
            "aladdin_user",
            "aladdin_passwd",
        ]
        missing = [v for v in required if not os.getenv(v)]
        if missing:
            if debug:
                print(
                    "  Missing environment variables for aladdin: "
                    + ", ".join(missing)
                    + "; skipping LLM call"
                )
            return None, None
    else:
        if debug:
            print(f"  Unsupported framework {FRAMEWORK}; skipping LLM call")
        return None, None

    payload = {"question": question_ctx, "sheets": sheets}
    if debug:
        print(f"  Prepared LLM payload with model '{model}'")

    try:
        res = _call_llm("xlsx_workbook_answer_slot.txt", payload, model=model)
    except Exception as exc:
        if debug:
            print(f"  LLM invocation failed: {exc}")
        return None, None

    if isinstance(res, dict):
        sheet, cell = res.get("sheet"), res.get("answer_cell")
        if debug:
            print(f"  LLM response: {res}")
        return sheet, cell

    if debug:
        print(f"  Unexpected LLM response type {type(res).__name__}: {res}")
    return None, None


def _llm_resolve_duplicate_slots(
    schema: List[Dict[str, Any]],
    sheets: Dict[str, Any],
    *,
    model: str,
    debug: bool = False,
) -> List[Dict[str, Any]]:
    """Ask the LLM to resolve answer-cell conflicts for questions.

    Parameters
    ----------
    schema:
        Current list of question entries with potential duplicate answer cells.
    sheets:
        Serialized workbook profiles to provide full context.
    model:
        Name of the LLM model to invoke.

    Returns
    -------
    List[Dict[str, Any]]
        Updated schema with revised ``sheet`` and ``answer_cell`` values.
    """

    payload = {
        "questions": [
            {
                "question_sheet": q["question_sheet"],
                "question_cell": q["question_cell"],
                "question_text": q.get("question_text"),
                "answer_sheet": q.get("sheet"),
                "answer_cell": q.get("answer_cell"),
            }
            for q in schema
        ],
        "sheets": sheets,
    }

    try:
        res = _call_llm("xlsx_resolve_duplicate_slots.txt", payload, model=model)
    except Exception as exc:  # pragma: no cover - defensive
        if debug:
            print(f"  Duplicate resolution failed: {exc}")
        return schema

    if debug:
        print(f"  Duplicate resolution LLM returned {len(res)} entries")

    mapping = {
        (r.get("question_sheet"), r.get("question_cell")): (
            r.get("sheet"),
            r.get("answer_cell"),
        )
        for r in res
    }

    for q in schema:
        key = (q["question_sheet"], q["question_cell"])
        if key in mapping:
            new_sheet, new_cell = mapping[key]
            if debug:
                print(f"  Reassigning {key}: {new_sheet}!{new_cell}")
            q["sheet"] = new_sheet or q["sheet"]
            q["answer_cell"] = new_cell

    return schema

# ---------------------------------------------------------------------------
# LLM driven pipeline
# ---------------------------------------------------------------------------


def profile_workbook(path: str, debug: bool = False) -> Dict[str, Any]:
    """Create a profile of every cell in the workbook.

    The profile captures basic style information so that subsequent LLM calls
    have rich context about how cells are formatted.  The structure returned is
    a ``dict`` mapping sheet names to a ``dict`` with ``max_row``, ``max_col``
    and a ``cells`` list containing the per-cell details.
    """

    if debug:
        print(f"Loading workbook for profiling: {path}")
    wb = load_workbook(path, data_only=True)
    profile: Dict[str, Any] = {}
    for ws in wb.worksheets:
        if debug:
            print(f"  Profiling sheet '{ws.title}'")
        cells: List[Dict[str, Any]] = []
        merged: set[Tuple[int, int]] = set()
        for rng in ws.merged_cells.ranges:
            merged.update({(r, c) for r, c in rng.cells})
        for row in ws.iter_rows():
            for cell in row:
                cells.append(
                    {
                        "row": cell.row,
                        "col": cell.column,
                        "value": cell.value,
                        "bold": bool(cell.font and cell.font.bold),
                        "italic": bool(cell.font and cell.font.italic),
                        "font": getattr(cell.font, "name", None),
                        "fill": _color_to_hex(cell.fill.start_color),
                        "border": {
                            "left": bool(cell.border.left.style),
                            "right": bool(cell.border.right.style),
                            "top": bool(cell.border.top.style),
                            "bottom": bool(cell.border.bottom.style),
                        },
                        "alignment": getattr(cell.alignment, "horizontal", None),
                        "locked": bool(getattr(cell.protection, "locked", False)),
                        "merged": (cell.row, cell.column) in merged,
                    }
                )
        profile[ws.title] = {
            "max_row": ws.max_row,
            "max_col": ws.max_column,
            "cells": cells,
        }
    if debug:
        print(f"Profiled workbook with {len(profile)} sheets")
    return profile


def _call_llm(prompt_file: str, payload: dict, *, model: str) -> Any:
    """Helper to invoke the selected LLM with a prompt template."""

    from answer_composer import CompletionsClient, get_openai_completion

    prompt_path = os.path.join(os.path.dirname(__file__), "prompts", prompt_file)
    with open(prompt_path, "r", encoding="utf-8") as f:
        template = f.read()
    prompt = template.replace("{{data}}", json.dumps(payload))

    if FRAMEWORK == "aladdin":
        resp = CompletionsClient(model=model).get_completion(prompt, json_output=True)
        content = resp[0] if isinstance(resp, tuple) else resp
    else:
        content, _ = get_openai_completion(prompt, model, json_output=True)
    return json.loads(content)


def _llm_macro_regions(
    profile: Dict[str, Any], *, model: str, debug: bool = False
) -> List[Dict[str, Any]]:
    """LLM step #1 – identify large rectangular regions in each sheet."""

    if debug:
        print(f"Identifying macro regions for {len(profile)} sheets")
    summaries = []
    for sheet, info in profile.items():
        summaries.append(
            {
                "sheet": sheet,
                "max_row": info["max_row"],
                "max_col": info["max_col"],
            }
        )
    try:
        res = _call_llm("xlsx_macro_regions.txt", summaries, model=model)
        if debug:
            print(f"  Macro region LLM returned {len(res)} regions")
        return res
    except Exception as exc:
        if debug:
            print(f"  Macro region detection failed: {exc}")
        return []


def _llm_zone_refinement(
    profile: Dict[str, Any], regions: List[Dict[str, Any]], *, model: str, debug: bool = False
) -> List[Dict[str, Any]]:
    """LLM step #2 – refine macro regions into potential answer zones."""

    if debug:
        print(f"Refining {len(regions)} regions into zones")
    zones: List[Dict[str, Any]] = []

    def refine(region: Dict[str, Any]) -> List[Dict[str, Any]]:
        sheet_profile = profile.get(region.get("sheet"), {})
        payload = {"region": region, "cells": sheet_profile.get("cells", [])}
        try:
            res = _call_llm("xlsx_zone_refinement.txt", payload, model=model)
            if debug:
                print(f"  Region {region.get('sheet')} -> {len(res)} zones")
            return res
        except Exception as exc:
            if debug:
                print(f"  Zone refinement failed for {region}: {exc}")
            return []

    with ThreadPoolExecutor() as ex:
        futures = [ex.submit(refine, region) for region in regions]
        for fut in as_completed(futures):
            zones.extend(fut.result())

    if debug:
        print(f"Total zones: {len(zones)}")
    return zones


def _llm_extract_candidates(
    profile: Dict[str, Any], zones: List[Dict[str, Any]], *, model: str, debug: bool = False
) -> List[Dict[str, Any]]:
    """LLM step #3 – from each zone extract candidate answer slots."""

    if debug:
        print(f"Extracting candidates from {len(zones)} zones")
    candidates: List[Dict[str, Any]] = []

    def extract(zone: Dict[str, Any]) -> List[Dict[str, Any]]:
        sheet_profile = profile.get(zone.get("sheet"), {})
        payload = {"zone": zone, "cells": sheet_profile.get("cells", [])}
        try:
            res = _call_llm("xlsx_slot_candidates.txt", payload, model=model)
            if debug:
                print(f"  Zone {zone.get('sheet')} -> {len(res)} candidates")
            return res
        except Exception as exc:
            if debug:
                print(f"  Candidate extraction failed for {zone}: {exc}")
            return []

    with ThreadPoolExecutor() as ex:
        futures = [ex.submit(extract, zone) for zone in zones]
        for fut in as_completed(futures):
            candidates.extend(fut.result())

    if debug:
        print(f"Total candidates: {len(candidates)}")
    return candidates


def _llm_score_and_assign(
    candidates: List[Dict[str, Any]], *, model: str, debug: bool = False
) -> List[Dict[str, Any]]:
    """LLM steps #4 and #5 – score candidates and pick winners."""

    if debug:
        print(f"Scoring {len(candidates)} candidates")
    if not candidates:
        if debug:
            print("No candidates to score")
        return []
    try:
        scored = _call_llm("xlsx_slot_scoring.txt", candidates, model=model)
        if debug:
            print(f"  Scoring step returned {len(scored)} entries")
    except Exception as exc:
        if debug:
            print(f"  Scoring failed: {exc}")
        return []

    # Index original candidates by ``slot_id`` so we can merge their richer
    # metadata (e.g. ``question_text`` or ``cell``) back into the scored
    # results.  The scoring prompt only returns a minimal subset of fields,
    # which previously caused that extra information to be dropped.
    by_slot: Dict[str, Dict[str, Any]] = {
        c.get("slot_id"): c for c in candidates if c.get("slot_id")
    }

    chosen: Dict[str, Dict[str, Any]] = {}
    for cand in scored:
        slot_id = cand.get("slot_id")
        merged = {**by_slot.get(slot_id, {}), **cand}
        # ``cell`` from the candidate is the answer location; expose it as
        # ``answer_cell`` unless already provided.
        if "cell" in merged and "answer_cell" not in merged:
            merged["answer_cell"] = merged.pop("cell")
        qid = merged.get("question_id") or merged.get("question_cell")
        best = chosen.get(qid)
        if debug:
            print(
                f"  Candidate {slot_id} for {qid}: score={merged.get('score')} cell={merged.get('answer_cell')}"
            )
            if best:
                print(
                    f"    Current best score={best.get('score')} cell={best.get('answer_cell')}"
                )
        if not best or merged.get("score", 0) > best.get("score", 0):
            if debug:
                action = "replacing best" if best else "initial best"
                print(f"    -> {action}")
            chosen[qid] = merged
    if debug:
        for qid, slot in chosen.items():
            print(
                f"  Final choice for {qid}: cell={slot.get('answer_cell')} score={slot.get('score')}"
            )
        print(f"Selected {len(chosen)} final slots")
    return list(chosen.values())


def extract_schema_from_xlsx(
    path: str,
    debug: bool = True,
    *,
    model: str = EXCEL_MODEL,
) -> List[Dict[str, Any]]:
    """Identify question cells in ``path`` and locate their answer slots via LLM.

    Each non-empty cell is examined with spaCy to determine whether it reads
    like a question or imperative.  For every such cell a context package is
    assembled consisting of the entire row and a 10×10 rectangle around the
    cell.  That context plus serialized profiles of **all** sheets are supplied
    to an LLM which chooses the best sheet and cell for the answer.  If the
    model is unavailable or declines to pick a location, ``answer_cell`` is
    ``None`` and the question sheet is used as a fallback.
    """

    wb = load_workbook(path, data_only=True)

    # Pre-serialize all sheets so we can present them to the model.
    sheet_profiles = {ws.title: _sheet_to_json(ws) for ws in wb.worksheets}

    tasks: List[Tuple[str, str, str, Dict[str, Any]]] = []

    for ws in wb.worksheets:
        if debug:
            print(f"Scanning sheet '{ws.title}'")
        for row in ws.iter_rows():
            for cell in row:
                value = cell.value
                if value is None:
                    continue
                text = str(value).strip()
                if not text:
                    continue
                if text.lower() in {"n/a", "na", "n.a.", "n a"}:
                    continue
                if not _spacy_is_question_or_imperative(text):
                    continue

                if debug:
                    print(f"  Question at {cell.coordinate}: {text}")
                qctx = _question_context(ws, cell)
                tasks.append((ws.title, cell.coordinate, text, qctx))

    def choose(task: Tuple[str, str, str, Dict[str, Any]]) -> Dict[str, Any]:
        sheet_title, coord, text, qctx = task
        answer_sheet, answer_cell = _llm_choose_answer_slot(
            qctx, sheet_profiles, model=model, debug=debug
        )
        if debug:
            print(
                f"    LLM chose {answer_sheet}!{answer_cell}" if answer_cell else "    LLM declined"
            )
        return {
            "sheet": answer_sheet or sheet_title,
            "question_sheet": sheet_title,
            "question_cell": coord,
            "question_text": text,
            "answer_cell": answer_cell,
        }

    with ThreadPoolExecutor() as ex:
        schema = list(ex.map(choose, tasks))

    # Detect duplicate answer cells and resolve them via a second LLM pass
    seen = {}
    duplicates = False
    for q in schema:
        cell = q.get("answer_cell")
        if not cell:
            continue
        key = (q.get("sheet"), cell)
        if key in seen:
            duplicates = True
            break
        seen[key] = True
    if duplicates:
        if debug:
            print("Duplicate answer cells detected; invoking resolution step")
        schema = _llm_resolve_duplicate_slots(schema, sheet_profiles, model=model, debug=debug)

    return schema


# Back‑compat alias so the CLI can import ask_sheet_schema from rfp
def ask_sheet_schema(xlsx_path: str, debug: bool = True) -> List[Dict[str, Any]]:
    """Compatibility wrapper for :func:`extract_schema_from_xlsx`.

    Parameters
    ----------
    xlsx_path:
        Path to the workbook to analyze.
    debug:
        Pass ``True`` to enable verbose debugging output (default).
    """
    return extract_schema_from_xlsx(xlsx_path, debug=debug)


def extract_slots_from_xlsx(path: str) -> Dict[str, Any]:
    """Extract text and formatting from an .xlsx file.

    The returned structure mirrors the DOCX slot finder in spirit but is
    simplified.  It returns a dictionary with ``doc_type`` set to
    ``"xlsx"`` and a ``sheets`` list, where each sheet contains a list of
    populated cells.  Each cell records its address (e.g. ``A1``), the
    text value, and basic formatting attributes (font color, bold,
    italic, background color, and border styles).
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    sheets: List[Dict[str, Any]] = []
    for ws in wb.worksheets:
        cells: List[Dict[str, Any]] = []
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue
                cell_info: Dict[str, Any] = {
                    "address": cell.coordinate,
                    "row": cell.row,
                    "column": get_column_letter(cell.column),
                    "value": str(cell.value),
                    "font_color": _color_to_hex(cell.font.color),
                    "bold": bool(cell.font.bold),
                    "italic": bool(cell.font.italic),
                    "bg_color": _color_to_hex(cell.fill.start_color),
                    "border": {
                        "left": cell.border.left.style,
                        "right": cell.border.right.style,
                        "top": cell.border.top.style,
                        "bottom": cell.border.bottom.style,
                    },
                }
                cells.append(cell_info)
        sheets.append({"name": ws.title, "cells": cells})
    return {"doc_type": "xlsx", "file": os.path.basename(path), "sheets": sheets}


def extract_cell_features(path: str) -> List[Dict[str, Any]]:
    """Flatten a workbook into a feature table for LLM classification.

    Each returned row has ``sheet``, ``row``, ``col``, ``cell`` address, the
    raw ``text`` value and a ``features`` dictionary containing simple style
    and text markers that may hint at questions.
    """
    wb = load_workbook(path, data_only=True)
    rows: List[Dict[str, Any]] = []
    interrogatives = (
        "who",
        "what",
        "when",
        "where",
        "why",
        "how",
        "do",
        "does",
        "can",
        "is",
        "are",
        "will",
        "should",
        "please",
    )
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue
                text = str(cell.value)
                low = text.strip().lower()
                features = {
                    "has_question_mark": "?" in text,
                    "starts_with_interrogative": bool(
                        re.match(rf"^({'|'.join(interrogatives)})\b", low)
                    ),
                    "bold": bool(cell.font and cell.font.bold),
                    "font_size": getattr(cell.font, "sz", None),
                    "fill_color": _color_to_hex(cell.fill.start_color),
                }
                rows.append(
                    {
                        "sheet": ws.title,
                        "row": cell.row,
                        "col": cell.column,
                        "cell": cell.coordinate,
                        "text": text,
                        "features": features,
                    }
                )
    return rows


def llm_classify_cells(
    cells: List[Dict[str, Any]], *, model: str = "gpt-4o-mini", batch_size: int = 8
) -> List[Dict[str, Any]]:
    """Classify cell schemas using an LLM.

    ``cells`` is expected to be the output of :func:`extract_cell_features`.
    The function batches the payload to reduce round trips.  Each batch is
    sent to an LLM prompt that labels the cells as ``QUESTION_HEADER``,
    ``QUESTION_BODY`` or ``NOT_QUESTION`` and returns a confidence score.
    """
    from answer_composer import get_openai_completion

    prompt_path = os.path.join(os.path.dirname(__file__), "prompts", "xlsx_classify_cells.txt")
    with open(prompt_path, "r", encoding="utf-8") as f:
        template = f.read()

    results: List[Dict[str, Any]] = []
    for i in range(0, len(cells), batch_size):
        batch = cells[i : i + batch_size]
        payload = json.dumps(
            [
                {"cell": c["cell"], "text": c["text"], "features": c["features"]}
                for c in batch
            ]
        )
        prompt = template.replace("{{cells}}", payload)
        content, _ = get_openai_completion(prompt, model, json_output=True)
        try:
            classified = json.loads(content)
        except Exception:
            continue
        results.extend(classified)

    # merge classifications back into the original structures
    by_cell = {r.get("cell"): r for r in results}
    for row in cells:
        r = by_cell.get(row["cell"])
        if r:
            row["label"] = r.get("label")
            row["confidence"] = r.get("confidence")
    return cells


def extract_questions_with_llm(path: str, model: str = "gpt-4o-mini") -> List[Dict[str, Any]]:
    """High level convenience wrapper combining feature extraction and LLM
    classification.

    The function returns rows labelled as question headers or bodies.  It does
    not attempt sophisticated block assembly but provides a foundation for
    downstream clustering.
    """
    table = extract_cell_features(path)
    classified = llm_classify_cells(table, model=model)
    return [row for row in classified if row.get("label") in {"QUESTION_HEADER", "QUESTION_BODY"}]


__all__ = [
    "extract_slots_from_xlsx",
    "extract_schema_from_xlsx",
    "ask_sheet_schema",
    "profile_workbook",
    "extract_cell_features",
    "llm_classify_cells",
    "extract_questions_with_llm",
]
